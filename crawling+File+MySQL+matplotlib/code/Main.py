from requests.packages import urllib3
import requests, json, openpyxl, os, time, pymysql
from openpyxl.utils import get_column_letter
import re
import pandas as pd
from sqlalchemy import create_engine
from scrapy.selector import Selector

def spider(code, proxy):
    urllib3.disable_warnings()
    title = ['地区', '年份', proxy]
    ans = []
    paramters = {'m': 'QueryData', 'dbcode': 'fsnd', 'rowcode': 'reg',
                 'colcode': 'sj',
                 'wds': '[{"wdcode":"zb","valuecode":"%s"}]' % code,
                 'dfwds': '[{"wdcode":"sj","valuecode": "2011-"}]',
                 }
    suffix = ''
    if proxy == '森林覆盖率' or proxy == '城市绿化率':
        suffix = '%'
    url = 'https://data.stats.gov.cn/easyquery.htm?cn=E0103'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3947.100 Safari/537.36'}
    html = requests.get(url, headers=headers, verify=False, params=paramters)
    res = json.loads(html.text)
    data = res['returndata']['datanodes']
    provinces = res['returndata']['wdnodes'][1]['nodes']
    j = 0
    for i in range(1, len(data)):
        if i % 12 == 0:
            j += 1
            continue
        ans.append([provinces[j]['cname'], data[i]['wds'][2]['valuecode'],
                    data[i]['data']['strdata'] + '%s' % suffix])
        print('地区', provinces[j]['cname'], '\t年份', data[i]['wds'][2]['valuecode'],
              '\t%s' % proxy,
              re.search(r'\d+\.?\d*',data[i]['data']['strdata']).group()
              + '%s' % suffix)
    return title, ans


def save_excel(title, data):
    path = '分省年度数据.xlsx'
    if os.path.exists(path):
        book = openpyxl.load_workbook(path)
        sheet_names = book.sheetnames
        sheet_name = '各省的{}'.format(title[2])
        if sheet_name not in sheet_names:
            sheet = book.create_sheet(sheet_name)
            sheet.append(title)
            for item in data:
                sheet.append(item)
            book.save(path)
            book.close()
        else:
            book.close()
    else:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = '各省的{}'.format(title[2])
        sheet.append(title)
        for item in data:
            sheet.append(item)
        book.save(path)
        book.close()


def adjust_column_width(sheetname,outputXlName):

    wb = openpyxl.load_workbook(outputXlName)
    ws = wb[sheetname]

    # 设置一个字典用于保存列宽数据
    dims = {}

    # 遍历表格数据，获取自适应列宽数据
    for row in ws.rows:
        for cell in row:
            if cell.value:
                # 遍历整个表格，把该列所有的单元格文本进行长度对比，找出最长的单元格
                # 在对比单元格文本时需要将中文字符识别为1.7个长度，英文字符识别为1个，这里只需要将文本长度直接加上中文字符数量即可
                # re.findall('([\u4e00-\u9fa5])', cell.value)能够识别大部分中文字符
                cell_len = 0.7 * len(re.findall('([\u4e00-\u9fa5])', str(cell.value))) + len(str(cell.value))
                dims[cell.column] = max((dims.get(cell.column, 0), cell_len))
    for col, value in dims.items():
        # 设置列宽，get_column_letter用于获取数字列号对应的字母列号，最后值+7是用来调整最终效果的
        ws.column_dimensions[get_column_letter(col)].width = value + 7

    wb.save(outputXlName)


def pre_save_mysql():
    path = '分省年度数据.xlsx'
    book = openpyxl.load_workbook(path)
    sheetnames = book.sheetnames
    for sheetname in sheetnames:
        sheet = book[sheetname]
        res = sheet.iter_rows(values_only=True)
        data = list(res)
        print(data)
        save_mysql(data)
    book.close()


def save_mysql(data):
    map = {'人均水资源量(立方米)': 'water', '森林覆盖率': 'forest',
           '二氧化硫排放量(万吨)': 'emit', '城市绿化率': 'green'}
    try:
        con = pymysql.Connect(host='localhost', port=3306,
                              user='root', passwd='123456', db='visualdata')
    except:
        con = pymysql.Connect(host='localhost', port=3306,
                              user='root', passwd='123456')
        cur = con.cursor()
        cur.execute('create database visualdata')
        cur.close()
    cur = con.cursor()
    cur.execute('use visualdata')
    tablename = map[data[0][2]]
    cur.execute('drop table if exists %s ' % tablename)
    print(tablename)
    ct = '''create table %s(
                region varchar(50),
                year int,
                value varchar(50)
            )
    ''' % tablename

    cur.execute(ct)
    for i in range(1, len(data)):
        cur.execute("insert into {} values('{}','{}','{}')"
                    .format(tablename, data[i][0], data[i][1], data[i][2]))
    cur.execute('select * from %s' % tablename)
    print(cur.fetchall())
    con.commit()
    cur.close()
    con.close()


def Multiple_Table_Connection():
    engine = create_engine('mysql+pymysql://root:123456'
                           '@localhost:3306/visualdata?charset=utf8')
    res = pd.read_sql('''select w.region 地区,w.year 年份,w.value 人均水资源量,
        f.value 森林覆盖率,e.value 二氧化硫排放量,g.value 城市绿化率
        from water w,forest f,emit e,green g 
        where w.region=f.region and w.year=f.year 
        and w.region=e.region and w.year=e.year 
        and w.region=g.region and w.year=g.year;
        ''', engine.connect())
    print(res)
    res.to_sql(name='link', con=engine.connect(), if_exists='replace')
    print('多表连接成功存入数据库')
    res.to_excel('多表连接后的全部数据.xlsx', sheet_name='全部数据', index=False)
    print('多表连接存入Excel成功')


def word_spider(url):
    headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/69.0.3947.100 Safari/537.36'}
    response=requests.get(url,headers=headers)
    response.encoding = response.apparent_encoding
    print(response.text)
    selector=Selector(response)
    res=selector.xpath('//div[@class="Custom_UnionStyle"]//div/text()').extract()
    print(res)
    fs=open('生态新闻.txt',mode='a+')
    for item in res:
        fs.writelines(item)
    fs.close()
    print('生态新闻爬取成功')

if __name__ == '__main__':
    print('--------------正在爬取，请稍后------------------')
    dic = {'人均水资源量': ["A0C0305", '人均水资源量(立方米)'], '森林覆盖率': ["A0C0A04", '森林覆盖率'],
           '二氧化硫排放量': ["A0C0601", '二氧化硫排放量(万吨)'], '城市绿化率': ['A0B0805', '城市绿化率']}

    for key in dic:
        quota = dic[key]
        title, ans = spider(quota[0], quota[1])
        save_excel(title, ans)
        print('数据已存入Excel文件')
        time.sleep(1)
    for key in dic:
        adjust_column_width('各省的' + dic[key][1],r'分省年度数据.xlsx')
    try:
        pre_save_mysql()
        Multiple_Table_Connection()
    except:
        print('数据库连接出错')
    adjust_column_width('全部数据',r'多表连接后的全部数据.xlsx')
    word_spider('https://www.mee.gov.cn/xxgk/hjyw/202304/t20230423_1027943.shtml')
    time.sleep(5)



